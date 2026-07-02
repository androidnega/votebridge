"""Parse CSV/Excel uploads for voter eligibility bulk import."""

from __future__ import annotations

import csv
import io
import re
from typing import Any

from django.core.files.uploadedfile import UploadedFile

MAX_IMPORT_ROWS = 5000

INDEX_HEADERS = {
    "index_number",
    "index number",
    "index",
    "index no",
    "index_no",
    "student index",
    "indexnumber",
}
EMAIL_HEADERS = {"email", "email address", "e-mail", "e_mail", "student email"}
ELIGIBLE_HEADERS = {"is_eligible", "eligible", "eligibility", "status"}
REASON_HEADERS = {"eligibility_reason", "reason", "note", "notes", "comment"}

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "y", "eligible"}:
        return True
    if normalized in {"0", "false", "no", "n", "ineligible", "not eligible"}:
        return False
    return default


def _map_headers(raw_headers: list[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, header in enumerate(raw_headers):
        normalized = _normalize_header(header)
        if not normalized:
            continue
        if normalized in INDEX_HEADERS:
            mapping[index] = "index_number"
        elif normalized in EMAIL_HEADERS:
            mapping[index] = "email"
        elif normalized in ELIGIBLE_HEADERS:
            mapping[index] = "is_eligible"
        elif normalized in REASON_HEADERS:
            mapping[index] = "eligibility_reason"
    return mapping


def _row_from_values(
    values: list[Any],
    header_mapping: dict[int, str],
    *,
    row_number: int,
    default_is_eligible: bool,
    default_reason: str,
) -> dict[str, Any] | None:
    row: dict[str, Any] = {"row_number": row_number}
    has_eligible_column = False

    for index, field in header_mapping.items():
        if index >= len(values):
            continue
        cell = values[index]
        if field == "index_number":
            row["index_number"] = str(cell).strip() if cell is not None else ""
        elif field == "email":
            row["email"] = str(cell).strip() if cell is not None else ""
        elif field == "is_eligible":
            has_eligible_column = True
            row["is_eligible"] = _parse_bool(cell, default_is_eligible)
        elif field == "eligibility_reason":
            text = str(cell).strip() if cell is not None else ""
            if text:
                row["eligibility_reason"] = text

    if not row.get("index_number") and not row.get("email"):
        return None

    if "is_eligible" not in row and not has_eligible_column:
        row["is_eligible"] = default_is_eligible
    if not row.get("eligibility_reason"):
        row["eligibility_reason"] = default_reason

    return row


def _parse_csv(
    content: bytes,
    *,
    default_is_eligible: bool,
    default_reason: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows: list[dict[str, Any]] = []
    errors: list[str] = []

    try:
        first_row = next(reader)
    except StopIteration:
        return [], ["The file is empty."]

    header_mapping = _map_headers(first_row)
    data_rows = reader

    if not header_mapping:
        # Single-column list without headers — treat first row as data too.
        candidate = str(first_row[0]).strip() if first_row else ""
        if candidate:
            rows.append(
                {
                    "row_number": 1,
                    "index_number": candidate,
                    "email": "",
                    "is_eligible": default_is_eligible,
                    "eligibility_reason": default_reason,
                }
            )
        data_rows = reader
        start_row_number = 2
    else:
        start_row_number = 2

    for offset, raw in enumerate(data_rows, start=start_row_number):
        if len(rows) >= MAX_IMPORT_ROWS:
            errors.append(f"Import limited to {MAX_IMPORT_ROWS} rows.")
            break
        parsed = _row_from_values(
            raw,
            header_mapping or {0: "index_number"},
            row_number=offset,
            default_is_eligible=default_is_eligible,
            default_reason=default_reason,
        )
        if parsed:
            rows.append(parsed)

    if not rows and not errors:
        errors.append("No voter rows found in the file.")

    return rows, errors


def _parse_xlsx(
    content: bytes,
    *,
    default_is_eligible: bool,
    default_reason: str,
) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel import is unavailable. Install openpyxl or upload CSV.") from exc

    workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    rows: list[dict[str, Any]] = []
    errors: list[str] = []

    try:
        first_row = next(iterator)
    except StopIteration:
        workbook.close()
        return [], ["The file is empty."]

    header_mapping = _map_headers(list(first_row))
    start_row_number = 2

    if not header_mapping:
        candidate = str(first_row[0]).strip() if first_row and first_row[0] is not None else ""
        if candidate:
            rows.append(
                {
                    "row_number": 1,
                    "index_number": candidate,
                    "email": "",
                    "is_eligible": default_is_eligible,
                    "eligibility_reason": default_reason,
                }
            )
        start_row_number = 2
    else:
        start_row_number = 2

    for offset, raw in enumerate(iterator, start=start_row_number):
        if len(rows) >= MAX_IMPORT_ROWS:
            errors.append(f"Import limited to {MAX_IMPORT_ROWS} rows.")
            break
        parsed = _row_from_values(
            list(raw),
            header_mapping or {0: "index_number"},
            row_number=offset,
            default_is_eligible=default_is_eligible,
            default_reason=default_reason,
        )
        if parsed:
            rows.append(parsed)

    workbook.close()

    if not rows and not errors:
        errors.append("No voter rows found in the file.")

    return rows, errors


def parse_eligibility_upload(
    uploaded_file: UploadedFile,
    *,
    default_is_eligible: bool = True,
    default_reason: str = "Bulk import",
) -> tuple[list[dict[str, Any]], list[str]]:
    filename = (uploaded_file.name or "").lower()
    extension = "." + filename.rsplit(".", 1)[-1] if "." in filename else ""

    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("Unsupported file type. Upload a CSV or Excel (.xlsx) file.")

    content = uploaded_file.read()
    if not content:
        return [], ["The file is empty."]

    if extension == ".csv":
        return _parse_csv(
            content,
            default_is_eligible=default_is_eligible,
            default_reason=default_reason,
        )

    return _parse_xlsx(
        content,
        default_is_eligible=default_is_eligible,
        default_reason=default_reason,
    )
